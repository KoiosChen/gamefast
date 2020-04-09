from openpyxl import load_workbook
from .. import db, logger
from ..MyModule.RequestPost import post_request
from ..models import MachineRoom, City, Contacts, machineroom_level, machineroom_type
from ..proccessing_data.proccess.public_methods import new_data_obj
import json
import uuid


def import_machine_room_file(filepath, sheet_name, start_row=1, columns=6):
    """

    :param filepath:
    :param sheet_name:
    :param start_row: start from 0
    :param columns: start from 0, the 6 here means totally have 7 columns
    :return:
    """
    wb = load_workbook(filepath)
    sheet = wb[sheet_name]

    row_title = [sheet['A1'].value, sheet['B1'].value, sheet['C1'].value, sheet['D1'].value, sheet['E1'].value,
                 sheet['F1'].value]

    target_title = ["city", "address", "level", "admin", "phone", "type", "lift", "room_name"]

    if not (set(row_title) - set(target_title)):
        # 如果表单模板标题与目标标题相同，则允许写入
        return [
            {'machine_room_name': row[0].value + machineroom_level[str(row[2].value)] + machineroom_type[
                str(row[5].value)] if str(row[7].value) == '0' else row[0].value + row[7].value,
             'machine_room_address': row[1].value,
             'machine_room_level': row[2].value,
             'machine_room_city': new_data_obj('City', **{'city': row[0].value}).id,
             'machine_room_type': row[5].value,
             'machine_room_admin': new_data_obj('Contacts', **{'name': row[3].value, 'phoneNumber': row[4].value}).id,
             "machine_room_lift": row[6].value}
            for index, row in enumerate(sheet.rows) if index >= start_row]
    else:
        return False


def import_device_file(filepath, sheet_name, start_row=1, columns=6):
    """

    :param filepath:
    :param sheet_name:
    :param start_row: start from 0
    :param columns: start from 0, the 6 here means totally has 7 columns
    :return:
    """
    wb = load_workbook(filepath)
    sheet = wb[sheet_name]

    row_title = [sheet['A1'].value, sheet['B1'].value, sheet['C1'].value, sheet['D1'].value, sheet['E1'].value,
                 sheet['F1'].value, sheet['G1'].value, sheet['H1'].value, sheet['I1'].value, sheet['J1'].value,
                 sheet['K1'].value]

    target_title = ["device_name", "ip", "device_owner", "username", "password", "enable_password", "login_method",
                    "monitor_status", "device_vendor", "device_model", "machine_room"]
    
    if not (set(row_title) - set(target_title)):
        # 如果表单模板标题与目标标题相同，则允许写入
        return [
            {'device_name': row[0].value if row[0].value else str(uuid.uuid4()),
             'device_ip': row[1].value,
             'device_owner': row[2].value,
             'username': row[3].value,
             'password': row[4].value,
             'enable_password': row[5].value,
             'login_method': row[6].value,
             'monitor_status': row[7].value,
             'device_vendor': row[8].value,
             'device_model': row[9].value,
             'machine_room': row[10].value
             } for index, row in enumerate(sheet.rows) if
            index >= start_row]
    else:
        return False


def import_cutover_file(filepath, sheet_name, start_row=1, columns=6):
    """

    :param filepath:
    :param sheet_name:
    :param start_row: start from 0
    :param columns: start from 0, the 6 here means totally has 7 columns
    :return:
    """
    wb = load_workbook(filepath)
    sheet = wb[sheet_name]
    return [{'customer_name': row[0].value, 'line_code': row[1].value, 'line_content': row[2].value,
             'protect': row[3].value} for index, row in enumerate(sheet.rows) if
            index >= start_row]
